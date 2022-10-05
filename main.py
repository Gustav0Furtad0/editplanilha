from statistics import quantiles
import openpyxl as opx;
#import pyexcel as opexcel;

arquivoFinal = opx.load_workbook(filename='itensBranetJF.xlsx')
planilhaFinal = arquivoFinal['MEDICAMENTOS']



def listaPlanilhaAxiliar():
    teste = "a"
    while teste != "NAO":
        arquivoAuxiliar = input('Digite o nome da planilha a ser inserida com seu formato: ')
        arquivoAuxiliar = opx.load_workbook(filename=arquivoAuxiliar)
        planilhaAuxiliar = arquivoAuxiliar["CMM UBS's"]
        procuraItens(planilhaAuxiliar)

def procuraItens(planilhaAuxiliar):
        i = 3
        while planilhaAuxiliar.cell(row=i, column=1).value != None:
            procura = int(planilhaAuxiliar.cell(row=i, column=1).value)
            for ix in range(3, 1050):
                if planilhaFinal.cell(row=ix, column=2).value != None:
                    if planilhaFinal.cell(row=ix, column=2).value.isdigit():
                        if int(planilhaFinal.cell(row=ix, column=2).value) == procura:
                            quantidade = planilhaAuxiliar.cell(row=i, column=5).value
                            planilhaFinal.cell(row=ix, column=8).value = quantidade
                            break           
            i += 1  
    

arquivoFinal.save(filename='testeresultado.xlsx')

print ('cabou-se')

# arquivofinal.save(filename='relatorios.xlsx')